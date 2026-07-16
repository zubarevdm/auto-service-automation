# -*- coding: utf-8 -*-
"""Импорт Excel-выгрузки из CRM. Колонки распознаются автоматически по названиям —
работает с выгрузками YCLIENTS, HelloClient, 1С и произвольными таблицами,
где строка = заказ-наряд (визит)."""
from pathlib import Path

from openpyxl import load_workbook

from . import db
from .normalize import norm_phone, norm_name, norm_date, norm_int, norm_amount

# Синонимы заголовков: ключ поля → подстроки, по которым узнаём колонку
COLUMN_HINTS = {
    "phone":    ("телефон", "тел.", "phone", "моб"),
    "name":     ("фио", "имя", "клиент", "заказчик", "name"),
    "brand":    ("марка", "brand"),
    "model":    ("модель", "model"),
    "car":      ("авто", "автомобиль", "машина", "тс"),
    "year":     ("год", "year"),
    "date":     ("дата", "date"),
    "amount":   ("сумма", "итог", "стоимост", "amount", "total", "оплач"),
    "mileage":  ("пробег", "одометр", "km"),
    "services": ("услуг", "работ", "наименован", "описан", "выполнен"),
}


def detect_columns(header_row) -> dict:
    """Сопоставляет заголовки выгрузки полям системы."""
    mapping = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        title = str(cell).strip().lower()
        for field, hints in COLUMN_HINTS.items():
            if field not in mapping and any(h in title for h in hints):
                mapping[field] = idx
                break
    return mapping


def split_car(raw) -> tuple[str, str]:
    """«Toyota Camry» → («Toyota», «Camry»)."""
    parts = str(raw or "").strip().split(None, 1)
    if not parts:
        return "", ""
    return parts[0], (parts[1] if len(parts) > 1 else "")


def import_file(conn, path: str | Path) -> dict:
    """Импортирует выгрузку. Возвращает статистику: строк, визитов, клиентов, брака."""
    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)

    mapping = {}
    header_seen = False
    stats = {"rows": 0, "new_visits": 0, "skipped": 0}

    for row in rows:
        if not header_seen:
            mapping = detect_columns(row)
            if "phone" in mapping and "date" in mapping:
                header_seen = True
            continue
        stats["rows"] += 1

        def col(field):
            i = mapping.get(field)
            return row[i] if i is not None and i < len(row) else None

        phone = norm_phone(col("phone"))
        visit_date = norm_date(col("date"))
        if not phone or not visit_date:
            stats["skipped"] += 1
            continue

        name = norm_name(col("name"))
        if "brand" in mapping:
            brand, model = str(col("brand") or "").strip(), str(col("model") or "").strip()
        else:
            brand, model = split_car(col("car"))

        client_id = db.upsert_client(conn, name, phone)
        car_id = db.upsert_car(conn, client_id, brand, model, norm_int(col("year")))
        is_new = db.insert_visit(
            conn, client_id, car_id, visit_date.isoformat(),
            norm_amount(col("amount")), norm_int(col("mileage")),
            str(col("services") or "").strip(),
        )
        if is_new:
            stats["new_visits"] += 1

    wb.close()
    if not header_seen:
        raise ValueError(
            "Не удалось распознать колонки: нужны хотя бы «Телефон» и «Дата». "
            f"Найдено: {mapping or 'ничего'}"
        )
    conn.commit()
    stats["clients"] = conn.execute("SELECT COUNT(*) c FROM clients").fetchone()["c"]
    stats["visits"] = conn.execute("SELECT COUNT(*) c FROM visits").fetchone()["c"]
    return stats
